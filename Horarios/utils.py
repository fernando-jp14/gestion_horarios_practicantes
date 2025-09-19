import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def export_horarios_to_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horarios de Recuperación"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados
    headers = ['Practicante', 'Días de Falta', 'Días de Recuperación']
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border

    # Datos
    for row_idx, horario in enumerate(queryset, start=2):
        dias_falta = ', '.join([d.nombre for d in horario.dias_falta.all()])
        dias_recuperacion = ', '.join([d.nombre for d in horario.dias_recuperacion.all()])
        row_data = [str(horario.practicante), dias_falta, dias_recuperacion]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = alignment_left
            cell.border = thin_border

    # Ajustar ancho columnas
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        adjusted_width = max_length + 4
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Congelar encabezado
    ws.freeze_panes = "A2"

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output